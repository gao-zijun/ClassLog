import os
import io
import random
import string
from datetime import datetime as dt
from collections import defaultdict
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFont
from . import models


def generate_password(length=8):
    """生成随机密码"""
    chars = string.ascii_letters + string.digits
    return ''.join(random.choice(chars) for _ in range(length))


def build_workbook():
    """
    构建完整的工作簿对象（包含所有班级工作表和三个汇总表），
    但不设置密码。密码设置在保存时进行。
    """
    wb = Workbook()
    # 删除默认工作表
    default_ws = wb.active
    wb.remove(default_ws)

    # 获取所有学生数据
    all_students = models.load_all_students()
    all_class_names = models.get_all_class_names()

    # 按班级分组记录
    class_records = defaultdict(list)
    for uname, info in all_students.items():
        if info.get('role') != 'student':
            continue
        cls = info.get('class', '未知班级')
        name = info.get('name', uname)
        for rec in info.get('records', []):
            time_str = rec.get('time', '')
            date_part = ''
            time_part = ''
            try:
                t = dt.strptime(time_str, "%Y-%m-%d %H:%M:%S")
                date_part = t.strftime("%Y-%m-%d")
                time_part = t.strftime("%H:%M:%S")
            except:
                time_part = rec.get('display_time', '')
                date_part = ''
            category = rec.get('category', '其他')
            if category not in ('奖励', '惩罚', '其他'):
                category = '其他'
            class_records[cls].append({
                'name': name,
                'date': date_part,
                'time': time_part,
                'category': category,
                'action': rec.get('action', '')
            })

    # ---------- 前三个汇总工作表 ----------
    reward_ws = wb.create_sheet(title="奖励记录学生")
    punish_ws = wb.create_sheet(title="惩罚记录学生")
    other_ws = wb.create_sheet(title="其他记录学生")

    reward_ws.append(["序号", "班级", "姓名", "日期", "时间", "记录"])
    punish_ws.append(["序号", "班级", "姓名", "日期", "时间", "记录"])
    other_ws.append(["序号", "班级", "姓名", "日期", "时间", "记录"])

    # 收集所有记录用于汇总
    all_records = []
    for cls, recs in class_records.items():
        for rec in recs:
            all_records.append({
                'class': cls,
                'name': rec['name'],
                'date': rec['date'],
                'time': rec['time'],
                'category': rec['category'],
                'action': rec['action']
            })

    sorted_all = sorted(all_records, key=lambda x: (x['date'], x['time']))
    reward_idx = 1
    punish_idx = 1
    other_idx = 1
    for rec in sorted_all:
        if rec['category'] == '奖励':
            reward_ws.append([reward_idx, rec['class'], rec['name'], rec['date'], rec['time'], rec['action']])
            reward_idx += 1
        elif rec['category'] == '惩罚':
            punish_ws.append([punish_idx, rec['class'], rec['name'], rec['date'], rec['time'], rec['action']])
            punish_idx += 1
        else:
            other_ws.append([other_idx, rec['class'], rec['name'], rec['date'], rec['time'], rec['action']])
            other_idx += 1

    # ---------- 班级工作表（排在汇总表后面） ----------
    sorted_classes = sorted(all_class_names)
    for cls in sorted_classes:
        ws = wb.create_sheet(title=str(cls)[:31])
        ws.append(["序号", "姓名", "日期", "时间", "记录类型", "记录"])
        records = sorted(class_records.get(cls, []), key=lambda x: (x['date'], x['time']))
        for i, rec in enumerate(records, 1):
            ws.append([i, rec['name'], rec['date'], rec['time'], rec['category'], rec['action']])

    return wb


def export_students_xlsx(password):
    """导出学生数据为加密的 xlsx 文件（BytesIO）"""
    wb = build_workbook()
    wb.security.workbookPassword = password
    wb.security.lockStructure = True
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _find_font():
    """查找可用的中文字体，返回字体路径，找不到返回 None"""
    candidates = [
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "C:/Windows/Fonts/simsun.ttc",
        "C:/Windows/Fonts/arial.ttf",
    ]
    for path in candidates:
        if os.path.exists(path):
            return path
    return None


def _worksheet_to_image(ws):
    """将 openpyxl 工作表渲染为 PNG 图片（BytesIO）"""
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([str(c) if c is not None else '' for c in row])
    if not data:
        return None

    num_cols = max(len(row) for row in data)
    for row in data:
        while len(row) < num_cols:
            row.append('')

    font_path = _find_font()
    header_font = ImageFont.truetype(font_path, 16) if font_path else ImageFont.load_default()
    body_font = ImageFont.truetype(font_path, 14) if font_path else ImageFont.load_default()
    title_font = ImageFont.truetype(font_path, 20) if font_path else ImageFont.load_default()

    # 计算列宽
    col_widths = []
    for c in range(num_cols):
        max_len = len(data[0][c]) if data else 0
        for row in data:
            max_len = max(max_len, len(row[c]))
        col_widths.append(max_len * 20 + 20)

    margin = 20
    header_height = 40
    row_height = 35
    title = ws.title or "Sheet"
    title_height = 50

    total_width = sum(col_widths) + margin * 2
    total_height = title_height + header_height + row_height * (len(data) - 1) + margin * 2

    image = Image.new('RGB', (total_width, total_height), 'white')
    draw = ImageDraw.Draw(image)

    # 标题
    title_bbox = draw.textbbox((0, 0), title, font=title_font)
    title_w = title_bbox[2] - title_bbox[0]
    draw.text(((total_width - title_w) / 2, margin), title, fill='black', font=title_font)

    # 表头
    y = margin + title_height
    x = margin
    for c in range(num_cols):
        draw.rectangle([x, y, x + col_widths[c], y + header_height], outline='black', width=1)
        text = data[0][c]
        bbox = draw.textbbox((0, 0), text, font=header_font)
        text_w = bbox[2] - bbox[0]
        text_x = x + (col_widths[c] - text_w) / 2
        draw.text((text_x, y + (header_height - (bbox[3] - bbox[1])) / 2), text, fill='black', font=header_font)
        x += col_widths[c]

    # 数据行
    for r in range(1, len(data)):
        y = margin + title_height + header_height + (r - 1) * row_height
        x = margin
        for c in range(num_cols):
            draw.rectangle([x, y, x + col_widths[c], y + row_height], outline='black', width=1)
            text = data[r][c]
            bbox = draw.textbbox((0, 0), text, font=body_font)
            text_w = bbox[2] - bbox[0]
            text_x = x + (col_widths[c] - text_w) / 2
            draw.text((text_x, y + (row_height - (bbox[3] - bbox[1])) / 2), text, fill='black', font=body_font)
            x += col_widths[c]

    img_buffer = io.BytesIO()
    image.save(img_buffer, format='PNG')
    img_buffer.seek(0)
    return img_buffer


def export_students_image_single(sheet_name):
    """导出指定工作表的图片"""
    wb = build_workbook()
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    return _worksheet_to_image(ws)


def export_students_images_zip(zip_password):
    """导出所有工作表图片，并打包成加密 ZIP"""
    import pyzipper
    wb = build_workbook()
    images = {}
    for name in wb.sheetnames:
        ws = wb[name]
        img_buffer = _worksheet_to_image(ws)
        if img_buffer:
            images[name + '.png'] = img_buffer.getvalue()

    zip_buffer = io.BytesIO()
    with pyzipper.AESZipFile(zip_buffer, 'w', compression=pyzipper.ZIP_DEFLATED, encryption=pyzipper.WZ_AES) as zf:
        zf.setpassword(zip_password.encode())
        for filename, data in images.items():
            zf.writestr(filename, data)
    zip_buffer.seek(0)
    return zip_buffer